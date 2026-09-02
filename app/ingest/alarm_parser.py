"""
app.ingest.alarm_parser —— 报警码记录的解析 / 归一化 / 校验，脚本导入、手工录入、Excel 导入共用。

- AlarmRecord：单条报警的不可变结构化记录（frozen、可 hash）。
- parse_record(dict)：归一化 → AlarmRecord；必填缺失抛 AlarmParseError。
- parse_jsonl / parse_excel：批量解析，坏行 yield BadRow（由调用方决定跳过/终止）。
- validate_record：返回错误消息 list（空 = OK）。
- build_embedding_text：向量化文本模板（[品牌][系统] 报警{code}{name} + 现象/原因/处置）。
- normalize_*：各字段独立归一化（便于单测 / 复用）。

归一化把 code_norm / brand / severity / category 收敛到合法枚举，与 kb.alarms CHECK 约束对齐。
"""

from __future__ import annotations

import json
import logging
import re
from collections.abc import Iterator
from dataclasses import dataclass, field, fields
from pathlib import Path
from typing import Any, Literal

logger = logging.getLogger(__name__)


# ===== 枚举常量（与 kb.alarms CHECK 约束对齐） =====

# 与 002_core_tables.sql 的 CHECK 约束一一对应
VALID_SEVERITIES: frozenset[str] = frozenset({"info", "warning", "fault", "fatal", "unknown"})
VALID_CATEGORIES: frozenset[str] = frozenset({
    "servo",        # 伺服
    "spindle",      # 主轴
    "pmc",          # PMC ladder
    "overtravel",   # 超程
    "program",      # 程序错误
    "system",       # 系统
    "estop",        # 急停
    "toolchange",   # 换刀
    "axis",         # 进给轴通用
    "hydraulic",    # 液压
    "pneumatic",    # 气动
    "encoder",      # 编码器 / 反馈
    "param",        # 参数非法
    "other",
    "unknown",
})
VALID_ORIGINS: frozenset[str] = frozenset({"ingest", "manual", "feedback"})

# 中 → 英 映射（数据源来自公开社区整理稿，常用中文表述）
_SEVERITY_ZH_MAP: dict[str, str] = {
    "信息": "info",
    "提示": "info",
    "通知": "info",
    "警告": "warning",
    "注意": "warning",
    "提醒": "warning",
    "报警": "warning",     # 中文"报警"泛指 → warning；具体级别看严重度
    "故障": "fault",
    "错误": "fault",
    "异常": "fault",
    "严重": "fault",
    "致命": "fatal",
    "急停": "fatal",
    "安全": "fatal",
}

_CATEGORY_ZH_MAP: dict[str, str] = {
    "伺服": "servo",
    "主轴": "spindle",
    "刀库": "toolchange",
    "换刀": "toolchange",
    "刀塔": "toolchange",
    "PMC": "pmc",
    "梯形": "pmc",
    "梯图": "pmc",
    "超程": "overtravel",
    "限位": "overtravel",
    "程序": "program",
    "编程": "program",
    "宏程序": "program",
    "系统": "system",
    "急停": "estop",
    "进给": "axis",
    "轴": "axis",
    "液压": "hydraulic",
    "气动": "pneumatic",
    "编码器": "encoder",
    "反馈": "encoder",
    "参数": "param",
    "参数非法": "param",
    "参数错误": "param",
}

# 品牌归一化（接受中文/英文/小写）
_BRAND_ALIAS_MAP: dict[str, str] = {
    "FANUC": "FANUC",
    "发那科": "FANUC",
    "FANUC ROBODRILL": "FANUC",
    "MITSUBISHI": "MITSUBISHI",
    "三菱": "MITSUBISHI",
    "MELDAS": "MITSUBISHI",
    "SIEMENS": "SIEMENS",
    "西门子": "SIEMENS",
    "SINUMERIK": "SIEMENS",
    "HEIDENHAIN": "HEIDENHAIN",
    "海德汉": "HEIDENHAIN",
}


# ===== 异常类型 =====

class AlarmParseError(ValueError):
    """单条记录解析失败。批量场景下让调用方决定 WARN+跳过 / 终止。"""


# ===== 不可变记录 =====

@dataclass(slots=True, frozen=True)
class AlarmRecord:
    """单条报警码的结构化记录。frozen 保证 hashable + 业务层安全共享。"""
    brand: str                              # FANUC / MITSUBISHI / SIEMENS ...
    code: str                               # 原始码（保留书写风格，如 'SV0401'）
    code_norm: str                          # 归一化码：UPPER + 去空白（用于唯一键）
    name: str                               # 报警名称
    controller: str = ""                    # 0i-MF / 31i / 828D ...
    category: str = "unknown"
    severity: str = "unknown"
    description: str = ""
    cause: str = ""
    action: str = ""
    safety_note: str = ""
    doc_id: int | None = None
    page_no: int | None = None
    origin: Literal["ingest", "manual", "feedback"] = "ingest"
    created_by: str | None = None
    extra: dict[str, Any] = field(default_factory=dict)


# ===== 规范化函数（pure，可单测） =====

def normalize_code(code: Any) -> str:
    """
    code_norm 规约：
    - 转 str 并 strip
    - 统一大写
    - **不去前导零**：SV0401 != SV401，破坏语义
    返回空串视为非法（让调用方走 validate 报错）。
    """
    if code is None:
        return ""
    s = re.sub(r"\s+", "", str(code)).upper()
    return s


def normalize_brand(brand: Any) -> str:
    """
    兼容 'fanuc' / '发那科' / 'FANUC ROBODRILL' 等写法 → 标准化品牌名。
    未知品牌按 'OTHER' 处理（不报错，方便用户先录入再调整）。
    """
    if brand is None:
        return ""
    raw = str(brand).strip().upper()
    if not raw:
        return ""
    if raw in _BRAND_ALIAS_MAP:
        return _BRAND_ALIAS_MAP[raw]
    # 尝试子串匹配（注意：空串会全中，必须先 short-circuit）
    for alias, canonical in _BRAND_ALIAS_MAP.items():
        if alias and (alias in raw or raw in alias):
            return canonical
    return raw  # 原样返回，让 validate 决定


def normalize_severity(value: Any) -> str:
    """
    严重度归一化 → VALID_SEVERITIES 之一。
    接受：英文小写（info/warning/fault/fatal/unknown）、英文大写、中文。
    未知 → 'unknown'。
    """
    if value is None or value == "":
        return "unknown"
    s = str(value).strip()
    if not s:
        return "unknown"
    low = s.lower()
    if low in VALID_SEVERITIES:
        return low
    if s in _SEVERITY_ZH_MAP:
        return _SEVERITY_ZH_MAP[s]
    # 兼容 "Error/Warn/Alarm" 这类英文单词
    word_map = {
        "error": "fault",
        "err": "fault",
        "alarm": "warning",
        "warn": "warning",
        "info": "info",
        "notice": "info",
        "fatal": "fatal",
        "critical": "fatal",
    }
    if low in word_map:
        return word_map[low]
    return "unknown"


def normalize_category(value: Any) -> str:
    """
    类别归一化 → VALID_CATEGORIES 之一。未知 → 'unknown'。
    """
    if value is None or value == "":
        return "unknown"
    s = str(value).strip()
    if not s:
        return "unknown"
    low = s.lower()
    if low in VALID_CATEGORIES:
        return low
    if s in _CATEGORY_ZH_MAP:
        return _CATEGORY_ZH_MAP[s]
    return "unknown"


def join_lines(value: Any) -> str:
    """list[str] / str / None → '\\n' 分隔的字符串。空值归一为空串。"""
    if value is None:
        return ""
    if isinstance(value, list):
        return "\n".join(str(v) for v in value if v is not None and str(v).strip())
    return str(value)


# ===== 解析 =====

# 必填字段
_REQUIRED_FIELDS = ("brand", "code", "name")


def parse_record(rec: dict[str, Any], *, origin: str = "ingest") -> AlarmRecord:
    """
    单条 dict → AlarmRecord。
    - 必填缺失 → AlarmParseError
    - 选填缺失或类型异常 → 静默用默认值（不阻塞整批）
    - code_norm 由 code 派生，不接受外部传入（避免不一致）
    - 不调用 embedding 生成；这是入库前的纯结构层
    """
    if not isinstance(rec, dict):
        raise AlarmParseError(f"record must be dict, got {type(rec).__name__}")

    missing = [f for f in _REQUIRED_FIELDS if not rec.get(f)]
    if missing:
        raise AlarmParseError(f"missing required fields {missing}")

    brand = normalize_brand(rec["brand"])
    code = str(rec["code"]).strip()
    code_norm = normalize_code(code)
    if not code_norm:
        raise AlarmParseError("code normalizes to empty string")

    name = str(rec["name"]).strip()
    if not name:
        raise AlarmParseError("name is empty after strip")

    if origin not in VALID_ORIGINS:
        raise AlarmParseError(f"invalid origin {origin!r}; must be one of {sorted(VALID_ORIGINS)}")

    return AlarmRecord(
        brand=brand,
        code=code,
        code_norm=code_norm,
        name=name,
        controller=str(rec.get("controller") or "").strip(),
        category=normalize_category(rec.get("category")),
        severity=normalize_severity(rec.get("severity")),
        description=str(rec.get("description") or "").strip(),
        cause=join_lines(rec.get("cause")),
        action=join_lines(rec.get("action")),
        safety_note=str(rec.get("safety_note") or "").strip(),
        doc_id=rec.get("doc_id"),
        page_no=rec.get("page_no"),
        origin=origin,  # type: ignore[arg-type]
        created_by=rec.get("created_by"),
        extra={
            k: v for k, v in rec.items()
            if k not in {
                "brand", "controller", "code", "category", "severity", "name",
                "description", "cause", "action", "safety_note",
                "doc_id", "page_no", "created_by",
            }
        },
    )


@dataclass(slots=True, frozen=True)
class BadRow:
    """批量解析时的坏行描述：行号 / 来源 / 原始记录 / 错误信息（供错误报表导出）。"""
    line_no: int
    source: str                # 文件名 / sheet 名
    raw: Any
    error: str


def parse_jsonl(path: Path, *, origin: str = "ingest") -> Iterator[AlarmRecord | BadRow]:
    """
    JSONL → 逐行解析。坏行 yield BadRow 而非抛异常，调用方可跳过继续。
    """
    if not path.exists():
        raise FileNotFoundError(f"jsonl not found: {path}")
    with path.open("r", encoding="utf-8") as f:
        for ln, line in enumerate(f, 1):
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
            except json.JSONDecodeError as e:
                yield BadRow(line_no=ln, source=path.name, raw=line, error=f"json decode: {e}")
                continue
            try:
                yield parse_record(obj, origin=origin)
            except AlarmParseError as e:
                yield BadRow(line_no=ln, source=path.name, raw=obj, error=str(e))


def parse_excel(
    path: Path,
    *,
    sheet_name: str | None = None,
    origin: str = "manual",
) -> Iterator[AlarmRecord | BadRow]:
    """
    Excel 解析（Excel 导入流程用）。第一行为字段名（key），第二行起为数据；
    与 parse_record 走同样归一化路径，坏行 yield BadRow。依赖 openpyxl。
    """
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ImportError("parse_excel requires openpyxl; install via requirements.txt") from e

    if not path.exists():
        raise FileNotFoundError(f"excel not found: {path}")

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    assert ws is not None

    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return

    header_list = [str(h).strip() if h is not None else "" for h in header]
    # 行号从 2 起（header 是第 1 行）
    for ln, row in enumerate(rows, 2):
        # 跳过空行
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        rec = dict(zip(header_list, row, strict=False))
        try:
            yield parse_record(rec, origin=origin)
        except AlarmParseError as e:
            yield BadRow(line_no=ln, source=ws.title, raw=rec, error=str(e))


# ===== 校验 =====

def validate_record(rec: AlarmRecord) -> list[str]:
    """
    业务校验：返回错误消息 list；空 list 表示 OK。
    故意宽松 —— 已经过了 parse_record 规范化，多数情况下 OK。
    严格约束保留给下游（DB CHECK / Excel 模板规则）。
    """
    errs: list[str] = []
    if not rec.brand:
        errs.append("brand is empty")
    if not rec.code:
        errs.append("code is empty")
    if not rec.code_norm:
        errs.append("code_norm is empty")
    if not rec.name:
        errs.append("name is empty")
    if rec.severity not in VALID_SEVERITIES:
        errs.append(f"severity {rec.severity!r} not in {sorted(VALID_SEVERITIES)}")
    if rec.category not in VALID_CATEGORIES:
        errs.append(f"category {rec.category!r} not in {sorted(VALID_CATEGORIES)}")
    if rec.origin not in VALID_ORIGINS:
        errs.append(f"origin {rec.origin!r} not in {sorted(VALID_ORIGINS)}")
    if rec.doc_id is not None and (not isinstance(rec.doc_id, int) or rec.doc_id <= 0):
        errs.append(f"doc_id must be positive int, got {rec.doc_id!r}")
    if rec.page_no is not None and (not isinstance(rec.page_no, int) or rec.page_no < 0):
        errs.append(f"page_no must be non-negative int, got {rec.page_no!r}")
    return errs


# ===== 向量化文本模板 =====

def build_embedding_text(rec: AlarmRecord) -> str:
    """
    构造用于向量化入库的文本，模板：
        [品牌][系统] 报警{code} {name}。
        现象：<description>
        原因：<cause 用换行>
        处置：<action 用换行>

    要点：字段间用换行分隔（切词不丢语义）；safety_note 不拼进 embedding
    （只在结果卡片与 LLM 上下文显示）；缺失字段整体省略而非填空串。
    """
    head = f"[{rec.brand}][{rec.controller}] 报警{rec.code} {rec.name}".strip()
    parts: list[str] = [head + "。"]
    if rec.description:
        parts.append(f"现象：{rec.description}")
    if rec.cause:
        parts.append(f"原因：\n{rec.cause}")
    if rec.action:
        parts.append(f"处置：\n{rec.action}")
    return "\n".join(parts)


# ===== 列出 AlarmRecord 字段名（便于 ORM-less 场景下做 INSERT 列展开） =====

def record_field_names() -> list[str]:
    return [f.name for f in fields(AlarmRecord) if f.name != "extra"]


def record_to_db_row(rec: AlarmRecord) -> dict[str, Any]:
    """
    AlarmRecord → 入库 dict（不含 extra）。
    业务层做 INSERT 时直接拿这个，省得手抄字段名。
    """
    return {f.name: getattr(rec, f.name) for f in fields(AlarmRecord) if f.name != "extra"}
