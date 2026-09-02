"""
app.api.knowledge —— 知识录入（手工/Excel 导入导出/文档上传）+ 条目与导入任务管理

POST   /entry                            手工录入报警码 / FAQ（含向量化）
PUT    /entry/{type}/{id}                编辑条目并重算向量
DELETE /entry/{type}/{id}                删除条目（级联清索引/向量）
POST   /entry/{type}/{id}/re-vectorize   重新生成向量
GET    /entries                          条目管理列表（alarm + faq 统一）
GET    /template?type=...                下载 Excel 模板
POST   /import/validate                  第一阶段：上传 + 校验（不写业务表）
POST   /import/{job_id}/confirm          第二阶段：实际入库（sync 或后台异步）
GET    /import/jobs                       导入历史列表
GET    /import/{job_id}                   导入进度
GET    /import/{job_id}/errors.xlsx      错误行报表
GET    /export?type=...                   按条件导出 xlsx
POST   /upload                            上传 md/pdf/txt 文档解析入库
GET    /documents                         文档列表 + 解析状态
GET    /documents/{doc_id}/chunks         文档在线查看（chunk 列表）
DELETE /documents/{doc_id}                删除文档（级联清 chunks + 源文件）
"""

from __future__ import annotations

import asyncio
import hashlib
import logging
import re
from dataclasses import replace
from datetime import UTC, datetime
from pathlib import Path as FilePath
from typing import Any, Literal

import psycopg
from fastapi import APIRouter, File, Form, HTTPException, Path, Query, Request, UploadFile
from fastapi.responses import Response
from pydantic import ValidationError

from app.config import get_settings
from app.db.repo.alarms import (
    delete_alarm_async,
    get_alarm_async,
    list_alarms_async,
    upsert_alarm_async,
    vectorize_one_alarm_async,
)
from app.db.repo.chunks import (
    delete_chunk_async,
    get_chunk_doc_id_async,
    get_chunk_with_title_async,
    insert_faq_async,
    list_faq_entries_async,
    update_chunk_text_async,
    vectorize_one_chunk_async,
)
from app.db.repo.documents import (
    delete_doc_async,
    get_doc_async,
    get_doc_by_hash_sync,
    insert_doc_pending_sync,
    list_chunks_async,
    list_documents_async,
    update_doc_status_async,
)
from app.db.repo.import_jobs import (
    get_job_async,
    insert_job_sync,
    list_jobs_async,
    update_progress_async,
)
from app.db.repo.kb_suggestions import reopen_suggestion_by_ref_async
from app.excel.export import export_alarms_to_xlsx, export_faq_to_xlsx
from app.excel.templates import generate_template_bytes
from app.excel.validate import parse_and_validate
from app.ingest.alarm_parser import (
    VALID_CATEGORIES,
    VALID_ORIGINS,
    VALID_SEVERITIES,
    AlarmParseError,
    build_embedding_text,
    parse_record,
    validate_record,
)
from app.ingest.loaders import load_any
from app.ingest.pipeline import ingest_manual
from app.ingest.vectorizer import text_from_alarm_row, vectorize_async
from app.llm.factory import build_embedding_provider
from app.schemas.entry import AlarmEntryRequest, EntryResponse, FAQEntryRequest

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/knowledge", tags=["knowledge"])

_VALID_ENTRY_ORIGINS = ("ingest", "manual", "feedback")


def _iso(dt: Any) -> str | None:
    return dt.isoformat() if isinstance(dt, datetime) else dt


async def _embed_single(text: str, request: Request) -> list[float]:
    """调 embedding Provider 嵌入单条文本"""
    cfg = request.app.state.cfg
    try:
        embedding = build_embedding_provider(cfg)
    except Exception as e:
        raise HTTPException(
            status_code=503,
            detail=f"embedding provider unavailable: {e}",
        ) from e
    vecs = await embedding.embed([text])
    if not vecs:
        raise HTTPException(status_code=502, detail="embedding returned empty")
    return vecs[0]


# ============ POST /api/knowledge/entry ============

@router.post("/entry", response_model=EntryResponse)
async def create_entry(payload: dict[str, Any], request: Request) -> EntryResponse:
    entry_type = payload.get("type")
    if entry_type == "alarm":
        return await _create_alarm(payload, request)
    elif entry_type == "faq":
        return await _create_faq(payload, request)
    raise HTTPException(
        status_code=422,
        detail=f"unknown entry type {entry_type!r}; expected 'alarm' or 'faq'",
    )


async def _create_alarm(
    payload: dict[str, Any], request: Request, origin: str = "manual",
) -> EntryResponse:
    try:
        req = AlarmEntryRequest(**payload)
    except ValidationError as e:
        raise HTTPException(status_code=422, detail=e.errors()) from e

    try:
        rec = parse_record(
            req.model_dump(exclude={"type", "created_by"}),
            origin=origin,
        )
        rec = replace(rec, created_by=req.created_by)
    except AlarmParseError as e:
        raise HTTPException(status_code=422, detail=str(e)) from e

    errs = validate_record(rec)
    if errs:
        raise HTTPException(status_code=422, detail=errs)
    if rec.category not in VALID_CATEGORIES:
        raise HTTPException(
            status_code=422,
            detail=f"category {rec.category!r} not in {sorted(VALID_CATEGORIES)}",
        )
    if rec.severity not in VALID_SEVERITIES:
        raise HTTPException(
            status_code=422,
            detail=f"severity {rec.severity!r} not in {sorted(VALID_SEVERITIES)}",
        )
    if rec.origin not in VALID_ORIGINS:
        raise HTTPException(
            status_code=422,
            detail=f"origin {rec.origin!r} not in {sorted(VALID_ORIGINS)}",
        )

    pool = request.app.state.pool
    new_id = await upsert_alarm_async(pool, rec)
    if new_id <= 0:
        raise HTTPException(status_code=500, detail="upsert_alarm failed")

    embed_text = build_embedding_text(rec)
    vector = await _embed_single(embed_text, request)
    ok = await vectorize_one_alarm_async(pool, new_id, vector)

    return EntryResponse(
        id=new_id,
        type="alarm",
        code_norm=rec.code_norm,
        vectorized=ok,
    )


async def _create_faq(payload: dict[str, Any], request: Request) -> EntryResponse:
    try:
        req = FAQEntryRequest(**payload)
    except ValidationError as e:
        raise HTTPException(status_code=422, detail=e.errors()) from e

    pool = request.app.state.pool
    doc_id, chunk_id = await insert_faq_async(
        pool,
        title=req.title,
        body=req.body,
        brand=req.brand,
        model_scope=req.model_scope,
        source=req.source,
        created_by=req.created_by,
    )
    if doc_id <= 0 or chunk_id <= 0:
        raise HTTPException(status_code=500, detail="insert FAQ failed")

    embed_text = f"{req.title}\n{req.body}".strip()
    vector = await _embed_single(embed_text, request)
    ok = await vectorize_one_chunk_async(pool, chunk_id, vector)

    return EntryResponse(
        id=chunk_id,
        type="faq",
        doc_id=doc_id,
        vectorized=ok,
    )


# ============ PUT /api/knowledge/entry/{type}/{id} ============

@router.put("/entry/{entry_type}/{entry_id}", response_model=EntryResponse)
async def update_entry(
    entry_type: str = Path(..., pattern="^(alarm|faq)$"),
    entry_id: int = Path(..., gt=0),
    payload: dict[str, Any] = ...,
    request: Request = ...,
) -> EntryResponse:
    pool = request.app.state.pool

    if entry_type == "faq":
        try:
            req = FAQEntryRequest(**payload)
        except ValidationError as e:
            raise HTTPException(status_code=422, detail=e.errors()) from e

        ok = await update_chunk_text_async(pool, entry_id, req.body)
        if not ok:
            raise HTTPException(status_code=404, detail=f"chunk id={entry_id} not found")

        embed_text = f"{req.title}\n{req.body}".strip()
        vector = await _embed_single(embed_text, request)
        await vectorize_one_chunk_async(pool, entry_id, vector)

        return EntryResponse(id=entry_id, type="faq", vectorized=True)

    elif entry_type == "alarm":
        payload2 = dict(payload, type="alarm")
        return await _create_alarm(payload2, request)

    raise HTTPException(status_code=422, detail=f"unknown entry type {entry_type!r}")


# ============ DELETE /api/knowledge/entry/{type}/{id} ============

@router.delete("/entry/{entry_type}/{entry_id}")
async def delete_entry(
    entry_type: str = Path(..., pattern="^(alarm|faq)$"),
    entry_id: int = Path(..., gt=0),
    request: Request = ...,
) -> dict[str, Any]:
    """删除条目。alarm 删 kb.alarms 行（tsv/embedding 同行随删）；faq 按 chunk
    删所属文档（FK 级联清 chunk + 向量），无文档则直接删 chunk。
    被待补充建议引用的条目，删除后建议置回 open。只删知识表，不动业务/日志表。"""
    pool = request.app.state.pool

    if entry_type == "alarm":
        ok = await delete_alarm_async(pool, entry_id)
        if not ok:
            raise HTTPException(status_code=404, detail=f"alarm id={entry_id} 不存在")
        await reopen_suggestion_by_ref_async(pool, {"type": "alarm", "id": entry_id})
        return {"deleted": entry_id, "type": "alarm"}

    # faq：FAQ 录入 = 一篇文档 + 1 个 chunk，删文档即可级联清理
    doc_id = await get_chunk_doc_id_async(pool, entry_id)
    if doc_id:
        ok = await delete_doc_async(pool, doc_id)
        if not ok:
            raise HTTPException(status_code=404, detail=f"chunk id={entry_id} 不存在")
    else:
        ok = await delete_chunk_async(pool, entry_id)
        if not ok:
            raise HTTPException(status_code=404, detail=f"chunk id={entry_id} 不存在")
    await reopen_suggestion_by_ref_async(pool, {"type": "faq", "chunk_id": entry_id})
    return {"deleted": entry_id, "type": "faq"}


# ============ POST /api/knowledge/entry/{type}/{id}/re-vectorize ============

@router.post("/entry/{entry_type}/{entry_id}/re-vectorize")
async def re_vectorize_entry(
    entry_type: str = Path(..., pattern="^(alarm|faq)$"),
    entry_id: int = Path(..., gt=0),
    request: Request = ...,
) -> dict[str, Any]:
    """从库内重建文本 → embed → 覆盖 embedding。向量不准或内容已改未刷新时用。"""
    pool = request.app.state.pool
    if entry_type == "alarm":
        row = await get_alarm_async(pool, entry_id)
        if not row:
            raise HTTPException(status_code=404, detail=f"alarm id={entry_id} 不存在")
        text = text_from_alarm_row(row)
        vector = await _embed_single(text, request)
        ok = await vectorize_one_alarm_async(pool, entry_id, vector)
        return {"id": entry_id, "type": "alarm", "vectorized": ok}

    ch = await get_chunk_with_title_async(pool, entry_id)
    if not ch:
        raise HTTPException(status_code=404, detail=f"faq chunk id={entry_id} 不存在")
    text = f"{ch['title']}\n{ch['content']}".strip()
    vector = await _embed_single(text, request)
    ok = await vectorize_one_chunk_async(pool, entry_id, vector)
    return {"id": entry_id, "type": "faq", "vectorized": ok}


# ============ GET /api/knowledge/entries（条目管理列表） ============

@router.get("/entries")
async def list_entries(
    request: Request,
    type: Literal["alarm", "faq"] | None = Query(None, description="alarm/faq/缺省=全部"),
    origin: str | None = Query(None, description="ingest/manual/feedback"),
    q: str | None = Query(
        None, max_length=64, description="关键字（报警码名/码 或 FAQ 标题/正文）",
    ),
    limit: int = Query(20, ge=1, le=200),
    offset: int = Query(0, ge=0),
) -> dict[str, Any]:
    """报警码 + FAQ 条目统一列表（条目管理页用）。
    alarm 的 id = kb.alarms.id；faq 的 id = kb.chunks.id（文档级在 doc_id）。"""
    pool = request.app.state.pool
    if origin and origin not in _VALID_ENTRY_ORIGINS:
        raise HTTPException(
            status_code=422, detail=f"origin 必须为 {list(_VALID_ENTRY_ORIGINS)} 之一",
        )

    def _alarm_item(r: dict) -> dict[str, Any]:
        return {
            "type": "alarm", "id": r["id"], "doc_id": None,
            "title": f"{r['brand']} {r['code_norm']} {r['name']}".strip(),
            "origin": r["origin"], "created_by": r["created_by"],
            "created_at": _iso(r["created_at"]), "vectorized": bool(r["vectorized"]),
        }

    def _faq_item(r: dict) -> dict[str, Any]:
        return {
            "type": "faq", "id": r["chunk_id"], "doc_id": r["doc_id"],
            "title": r["title"], "origin": r["origin"],
            "created_by": r["created_by"], "created_at": _iso(r["created_at"]),
            "vectorized": bool(r["vectorized"]),
        }

    if type == "alarm":
        rows, total = await list_alarms_async(
            pool, q=q, origin=origin, limit=limit, offset=offset,
        )
        return {
            "total": total, "items": [_alarm_item(r) for r in rows],
            "limit": limit, "offset": offset,
        }

    if type == "faq":
        rows, total = await list_faq_entries_async(
            pool, q=q, origin=origin, limit=limit, offset=offset,
        )
        return {
            "total": total, "items": [_faq_item(r) for r in rows],
            "limit": limit, "offset": offset,
        }

    # 全部：两表各取足量后合并，按 created_at 倒序再分页（数据量级万级以内够用）
    big = max(limit, 1000)
    alarms, _at = await list_alarms_async(pool, q=q, origin=origin, limit=big, offset=0)
    faqs, _ft = await list_faq_entries_async(pool, q=q, origin=origin, limit=big, offset=0)
    merged = [_alarm_item(r) for r in alarms] + [_faq_item(r) for r in faqs]
    merged.sort(key=lambda x: x["created_at"] or "", reverse=True)
    total = len(merged)
    return {
        "total": total, "items": merged[offset:offset + limit],
        "limit": limit, "offset": offset,
    }


# ============ GET /api/knowledge/template ============

@router.get("/template")
async def get_template(
    type: Literal["alarm", "faq", "machine", "maintenance"] = Query(..., description="模板类型"),
) -> Response:
    """下载 Excel 模板（含说明行 + 枚举下拉）"""
    try:
        content = generate_template_bytes(type)
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e)) from e
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{type}_template.xlsx"',
        },
    )


# ============ POST /api/knowledge/import/validate ============

@router.post("/import/validate")
async def import_validate(
    request: Request,
    type: Literal["alarm", "faq", "machine", "maintenance"] = Query(...),
    dup_strategy: Literal["skip", "overwrite", "duplicate"] = Query("skip"),
    file: UploadFile = File(...),
) -> dict[str, Any]:
    """第一阶段：上传 + 校验，不写业务表（仅落一条 kb.import_jobs status='previewing'）。
    返回 job_id/total_rows/valid_rows/dup_rows/error_rows/sample_errors（前 50）/dup_strategy。"""
    content = await file.read()
    if not content:
        raise HTTPException(status_code=422, detail="empty file")

    file_hash = hashlib.sha256(content).hexdigest()
    filename = file.filename or f"upload_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    # 1) 解析 + 校验
    validation = parse_and_validate(content, type)

    # 2) 对 alarm 类型做 dup 检查（命中 unique key 的行号集合）
    dup_lines: set[int] = set()
    if type == "alarm":
        cfg = get_settings()
        try:
            with psycopg.connect(**cfg.db_dsn_kwargs()) as conn, conn.cursor() as cur:
                # 收集所有 (brand, code_norm) 对
                pairs: list[tuple[str, str]] = []
                for r in validation.parsed_rows:
                    b = (r.get("brand") or "").upper().strip()
                    c = (r.get("code") or "").upper().strip()
                    if b and c:
                        pairs.append((b, c))
                if pairs:
                    placeholders = ",".join(["(%s, %s)"] * len(pairs))
                    flat_params: list[Any] = []
                    for b, c in pairs:
                        flat_params.extend([b, c])
                    cur.execute(
                        f"SELECT brand, code_norm FROM kb.alarms "
                        f"WHERE (brand, code_norm) IN ({placeholders})",
                        flat_params,
                    )
                    existing = {(b, c) for b, c in cur.fetchall()}
                    for r in validation.parsed_rows:
                        b = (r.get("brand") or "").upper().strip()
                        c = (r.get("code") or "").upper().strip()
                        if (b, c) in existing:
                            dup_lines.add(r.get("__line_no__", -1))
        except Exception as e:
            logger.warning("[import/validate] dup check failed: %s", e)
            # dup 检查失败不影响 validate 结果，让 confirm 阶段去重

    dup_rows_count = len(dup_lines)
    valid_rows_count = validation.valid_rows - dup_rows_count
    error_rows_count = validation.error_rows

    # 3) INSERT 一条 previewing 记录到 kb.import_jobs
    cfg = get_settings()
    errors_payload = [
        {"row": e.line_no, "field": e.field, "reason": e.reason}
        for e in validation.errors
    ]
    with psycopg.connect(**cfg.db_dsn_kwargs()) as conn:
        job_id = insert_job_sync(
            conn,
            job_type=type,
            filename=filename,
            file_hash=file_hash,
            total_rows=validation.total_rows,
            valid_rows=max(0, valid_rows_count),
            dup_rows=dup_rows_count,
            error_rows=error_rows_count,
            dup_strategy=dup_strategy,
            status="previewing",
            errors=errors_payload,
            created_by=None,
        )

    sample_errors = [
        {"row": e.line_no, "field": e.field, "reason": e.reason}
        for e in validation.errors[:50]
    ]
    return {
        "job_id": job_id,
        "total_rows": validation.total_rows,
        "valid_rows": max(0, valid_rows_count),
        "dup_rows": dup_rows_count,
        "error_rows": error_rows_count,
        "sample_errors": sample_errors,
        "dup_strategy": dup_strategy,
        "dup_lines_lines": sorted(dup_lines),
    }


# ============ POST /api/knowledge/import/{job_id}/confirm ============

@router.post("/import/{job_id}/confirm")
async def import_confirm(
    request: Request,
    job_id: int = Path(..., gt=0),
    sync: bool = Query(False, description="True=同步等待；False=后台异步"),
) -> dict[str, Any]:
    """
    第二阶段：实际入库。
    sync=True 等待导入完成（测试用）；sync=False 立即返回 job_id，后台异步进度。
    """
    pool = request.app.state.pool
    cfg = request.app.state.cfg

    # 1) 读取 job
    job = await get_job_async(pool, job_id)
    if not job:
        raise HTTPException(status_code=404, detail=f"job_id {job_id} not found")
    if job["status"] != "previewing":
        raise HTTPException(
            status_code=409,
            detail=f"job_id {job_id} already in status {job['status']!r}, cannot confirm",
        )

    # 2) job 不存原文件：前端需把 /import/validate 的合法 rows 在 body.rows 里原样再传
    body = await request.json() if request.headers.get("content-type") == "application/json" else {}
    rows = body.get("rows")
    if not rows:
        raise HTTPException(
            status_code=422,
            detail="missing 'rows' in request body; front-end should "
                   "echo back the validated rows from /import/validate",
        )

    dup_strategy: str = job["dup_strategy"]
    job_type: str = job["job_type"]

    if sync:
        # 同步执行（测试场景，立即可见结果）
        result = await _execute_import(
            pool, cfg, job_type, rows, dup_strategy, job_id,
        )
        return {"job_id": job_id, "sync": True, **result}

    # 异步：fire-and-forget，后台跑
    asyncio_task = asyncio.create_task(  # noqa: F841
        _execute_import(pool, cfg, job_type, rows, dup_strategy, job_id)
    )
    return {"job_id": job_id, "sync": False, "started": True}


async def _execute_import(
    pool,
    cfg,
    job_type: str,
    rows: list[dict[str, Any]],
    dup_strategy: str,
    job_id: int,
) -> dict[str, Any]:
    """实际入库逻辑（同步/异步都走这条）"""
    await update_progress_async(pool, job_id, status="importing")
    imported = 0
    vectorized = 0
    failed = 0
    failed_errors: list[dict[str, Any]] = []

    if job_type == "alarm":
        for r in rows:
            try:
                rec = parse_record(
                    {k: v for k, v in r.items() if not k.startswith("__")},
                    origin="ingest",
                )
                new_id = await upsert_alarm_async(pool, rec)
                imported += 1
                # 向量化
                embed_text = build_embedding_text(rec)
                try:
                    embedding = build_embedding_provider(cfg)
                    vecs = await embedding.embed([embed_text])
                    if vecs and vecs[0]:
                        ok = await vectorize_one_alarm_async(pool, new_id, vecs[0])
                        if ok:
                            vectorized += 1
                except Exception as e:
                    logger.warning("[import] embed failed for alarm id=%d: %s", new_id, e)
            except Exception as e:
                failed += 1
                failed_errors.append({
                    "row": r.get("__line_no__", -1),
                    "field": None,
                    "reason": str(e),
                })

    elif job_type == "faq":
        for r in rows:
            try:
                title = r.get("title") or ""
                body = r.get("body") or ""
                if not title or not body:
                    failed += 1
                    failed_errors.append({
                        "row": r.get("__line_no__", -1),
                        "field": "title/body",
                        "reason": "必填为空",
                    })
                    continue
                doc_id, chunk_id = await insert_faq_async(
                    pool,
                    title=title,
                    body=body,
                    brand=r.get("brand"),
                    model_scope=[
                        s.strip()
                        for s in (r.get("model_scope") or "").split(",")
                        if s.strip()
                    ] or None,
                    source=r.get("source"),
                    created_by=None,
                )
                imported += 1
                try:
                    embedding = build_embedding_provider(cfg)
                    vecs = await embedding.embed([f"{title}\n{body}"])
                    if vecs and vecs[0]:
                        await vectorize_one_chunk_async(pool, chunk_id, vecs[0])
                        vectorized += 1
                except Exception as e:
                    logger.warning("[import] embed failed for faq chunk id=%d: %s", chunk_id, e)
            except Exception as e:
                failed += 1
                failed_errors.append({
                    "row": r.get("__line_no__", -1),
                    "field": None,
                    "reason": str(e),
                })

    else:
        # machine / maintenance：本批只做报警码 + FAQ 两种；其它类型直接报错
        failed = len(rows)
        failed_errors.append({
            "row": -1,
            "field": None,
            "reason": f"import type {job_type!r} not yet supported in W2.10",
        })

    finished_at = datetime.now(UTC)
    await update_progress_async(
        pool, job_id,
        imported_rows=imported,
        vectorized=vectorized,
        status="done" if failed == 0 else "done",   # 部分失败也算 done，errors 体现
        finished_at=finished_at,
        errors=(failed_errors or None),
    )

    return {
        "imported": imported,
        "vectorized": vectorized,
        "failed": failed,
        "failed_errors": failed_errors,
    }


# ===== 导入历史列表 =====
# 注意：必须在 /import/{job_id} 之前声明（FastAPI 按声明顺序匹配）

@router.get("/import/jobs")
async def list_import_jobs(
    request: Request,
    limit: int = Query(default=20, ge=1, le=100),
    offset: int = Query(default=0, ge=0),
) -> dict[str, Any]:
    """Excel 批量导入任务分页列表（id 倒序），供 KnowledgeView 显示历史"""
    pool = request.app.state.pool
    jobs, total = await list_jobs_async(pool, limit=limit, offset=offset)
    return {"total": total, "items": jobs, "limit": limit, "offset": offset}


# ============ GET /api/knowledge/import/{job_id} ============

@router.get("/import/{job_id}")
async def import_status(
    request: Request,
    job_id: int = Path(..., gt=0),
) -> dict[str, Any]:
    """进度查询"""
    pool = request.app.state.pool
    job = await get_job_async(pool, job_id)
    if not job:
        raise HTTPException(status_code=404, detail=f"job_id {job_id} not found")
    # datetime 转 ISO 字符串方便前端
    for k in ("created_at", "finished_at"):
        v = job.get(k)
        if isinstance(v, datetime):
            job[k] = v.isoformat()
    return job


# ============ GET /api/knowledge/import/{job_id}/errors.xlsx ============

@router.get("/import/{job_id}/errors.xlsx")
async def import_errors_xlsx(
    request: Request,
    job_id: int = Path(..., gt=0),
) -> Response:
    """错误行报表 xlsx（原行 + 新增"错误原因"列）"""
    pool = request.app.state.pool
    job = await get_job_async(pool, job_id)
    if not job:
        raise HTTPException(status_code=404, detail=f"job_id {job_id} not found")

    errors = job.get("errors") or []
    if not isinstance(errors, list):
        errors = []

    # 构造 xlsx
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "errors"
    ws.cell(row=1, column=1, value="row")
    ws.cell(row=1, column=2, value="field")
    ws.cell(row=1, column=3, value="reason")
    for i, e in enumerate(errors, start=2):
        ws.cell(row=i, column=1, value=e.get("row") if isinstance(e, dict) else None)
        ws.cell(row=i, column=2, value=e.get("field") if isinstance(e, dict) else None)
        ws.cell(row=i, column=3, value=e.get("reason") if isinstance(e, dict) else None)

    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)

    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="import_{job_id}_errors.xlsx"'},
    )


# ============ GET /api/knowledge/export ============

@router.get("/export")
async def export_data(
    type: Literal["alarm", "faq"] = Query(...),
    brand: str | None = Query(None),
    origin: str | None = Query(None),
) -> Response:
    """按条件导出 xlsx，列结构与模板一致"""
    if type == "alarm":
        content = export_alarms_to_xlsx(brand=brand, origin=origin)
    elif type == "faq":
        content = export_faq_to_xlsx(brand=brand)
    else:
        raise HTTPException(status_code=422, detail=f"unsupported type {type!r}")

    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{type}_export.xlsx"',
        },
    )


# ============ 文档上传 / 列表 / 删除 ============

_UPLOAD_DOC_TYPES = ("manual", "alarm_table", "maintenance_std", "sop", "faq", "other")
_UPLOAD_SUFFIXES = (".md", ".markdown", ".pdf", ".txt")


def _project_root() -> FilePath:
    """app/api/knowledge.py → 项目根目录（data/ 的父级）"""
    return FilePath(__file__).resolve().parent.parent.parent


def _upload_dir() -> FilePath:
    return _project_root() / "data" / "uploaded"


async def _process_upload(
    pool,
    cfg,
    *,
    doc_id: int,
    file_path: FilePath,
    title: str,
    doc_type: str,
    brand: str | None,
    model_scope: list[str],
    created_by: str | None,
) -> dict[str, Any]:
    """后台解析 doc：pending → parsing → ready/failed。
    ingest 异常 → failed + error_msg；仅向量化失败 → 告警不标废（内容已入库，
    可走 embedding IS NULL 断点续传补跑）。"""
    try:
        await update_doc_status_async(pool, doc_id, "parsing")
        pages = load_any(file_path)
        if not pages or not any(p.text.strip() for p in pages):
            raise ValueError("文件解析结果为空（无法提取任何文本）")

        with psycopg.connect(**cfg.db_dsn_kwargs()) as conn:
            result = ingest_manual(
                conn,
                title=title,
                pages=pages,
                doc_type=doc_type,
                brand=brand,
                model_scope=model_scope,
                source_file=str(file_path.relative_to(_project_root())),
                file_path=file_path,
                existing_doc_id=doc_id,
                created_by=created_by,
            )

        try:
            provider = build_embedding_provider(cfg)
            vr = await vectorize_async(pool, "chunks", provider)
            vectorized = vr.embedded
        except Exception as ve:  # noqa: BLE001
            logger.warning("[upload] doc_id=%d 向量化失败（内容已入库）: %s", doc_id, ve)
            vectorized = 0

        return {
            "status": "ready",
            "parents": result.parent_count,
            "children": result.child_count,
            "vectorized": vectorized,
        }
    except Exception as e:  # noqa: BLE001
        logger.exception("[upload] doc_id=%d 解析失败", doc_id)
        await update_doc_status_async(pool, doc_id, "failed", error_msg=str(e))
        return {"status": "failed", "error": str(e)}


@router.post("/upload")
async def upload_document(
    request: Request,
    file: UploadFile = File(...),
    title: str | None = Form(None, description="文档标题，缺省用文件名"),
    doc_type: str = Form("manual", description="manual/maintenance_std/sop/alarm_table/other"),
    brand: str | None = Form(None, description="适用品牌，如 FANUC"),
    model_scope: str | None = Form(None, description="适用机型，逗号分隔，如 VMC850,TC500"),
    created_by: str | None = Form(None, description="工号"),
    sync: bool = Query(False, description="True=同步解析后返回；False=后台异步（默认）"),
) -> dict[str, Any]:
    """
    上传 md/pdf/txt 文档，默认后台异步解析入库（sync=true 则同步等待返回）。
    校验类型与后缀 → file_hash 重复检测（重复 409）→ 存 data/uploaded/ → 解析入库。
    Excel 报警码/FAQ 走 /import，本端点不支持。
    """
    content = await file.read()
    if not content:
        raise HTTPException(status_code=422, detail="empty file")

    filename = file.filename or "untitled"
    suffix = FilePath(filename).suffix.lower()
    if suffix not in _UPLOAD_SUFFIXES:
        raise HTTPException(
            status_code=422,
            detail=(
                f"不支持的文件类型 {suffix!r}；支持 md/pdf/txt"
                "（Excel 走 /api/knowledge/import）"
            ),
        )
    if doc_type not in _UPLOAD_DOC_TYPES:
        detail = f"doc_type 非法；支持 {list(_UPLOAD_DOC_TYPES)}"
        raise HTTPException(status_code=422, detail=detail)

    cfg = request.app.state.cfg
    pool = request.app.state.pool
    file_hash = hashlib.sha256(content).hexdigest()
    doc_title = (title or "").strip() or FilePath(filename).stem

    # 1) 重复检测（file_hash UNIQUE）
    with psycopg.connect(**cfg.db_dsn_kwargs()) as conn:
        dup = get_doc_by_hash_sync(conn, file_hash)
    if dup:
        raise HTTPException(
            status_code=409,
            detail=f"文件已存在（doc_id={dup['id']}，status={dup['status']}）",
        )

    # 2) 落盘 data/uploaded/
    upload_dir = _upload_dir()
    upload_dir.mkdir(parents=True, exist_ok=True)
    safe_name = re.sub(r"[^A-Za-z0-9._-]", "_", FilePath(filename).name)
    stamp = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")
    abs_path = upload_dir / f"{stamp}_{safe_name}"
    abs_path.write_bytes(content)

    # 3) INSERT pending doc
    model_scope_list = [s.strip() for s in (model_scope or "").split(",") if s.strip()]
    with psycopg.connect(**cfg.db_dsn_kwargs()) as conn:
        doc_id = insert_doc_pending_sync(
            conn,
            title=doc_title,
            doc_type=doc_type,
            brand=brand,
            model_scope=model_scope_list,
            source_file=str(abs_path.relative_to(_project_root())),
            file_hash=file_hash,
            created_by=created_by,
        )
    if doc_id <= 0:
        raise HTTPException(status_code=500, detail="insert document failed")

    # 4) 解析（同步等待 or 后台异步）
    if sync:
        result = await _process_upload(
            pool, cfg, doc_id=doc_id, file_path=abs_path, title=doc_title,
            doc_type=doc_type, brand=brand, model_scope=model_scope_list,
            created_by=created_by,
        )
        return {"doc_id": doc_id, "sync": True, **result}

    asyncio.create_task(_process_upload(  # noqa: RUF006 —— 后台 fire-and-forget（V1）
        pool, cfg, doc_id=doc_id, file_path=abs_path, title=doc_title,
        doc_type=doc_type, brand=brand, model_scope=model_scope_list,
        created_by=created_by,
    ))
    return {"doc_id": doc_id, "sync": False, "status": "parsing"}


@router.get("/documents")
async def list_documents(
    request: Request,
    status: str | None = Query(None, description="过滤解析状态 pending/parsing/ready/failed"),
    limit: int = Query(20, ge=1, le=200),
    offset: int = Query(0, ge=0),
) -> dict[str, Any]:
    """文档列表 + 解析状态 + chunk 数（KnowledgeView 轮询解析进度用）"""
    pool = request.app.state.pool
    items, total = await list_documents_async(pool, status=status, limit=limit, offset=offset)
    return {"total": total, "items": items}


# ===== 文档在线查看（chunks 列表）=====
# 必须在 /documents/{doc_id} 之前声明（FastAPI 按声明顺序匹配）

@router.get("/documents/{doc_id}/chunks")
async def list_doc_chunks(
    request: Request,
    doc_id: int,
    limit: int = Query(default=50, ge=1, le=200),
    offset: int = Query(default=0, ge=0),
) -> dict[str, Any]:
    """列出指定文档的所有 chunks（在线查看文档内容用）"""
    pool = request.app.state.pool
    doc = await get_doc_async(pool, doc_id)
    if not doc:
        raise HTTPException(status_code=404, detail=f"document id={doc_id} 不存在")
    items, total = await list_chunks_async(pool, doc_id, limit=limit, offset=offset)
    return {
        "doc_id": doc_id,
        "title": doc["title"],
        "total": total,
        "items": items,
        "limit": limit,
        "offset": offset,
    }


@router.delete("/documents/{doc_id}")
async def delete_document(
    request: Request,
    doc_id: int = Path(..., gt=0),
) -> dict[str, Any]:
    """删除文档（级联删 chunks）+ 清理已存源文件"""
    pool = request.app.state.pool
    doc = await get_doc_async(pool, doc_id)
    if not doc:
        raise HTTPException(status_code=404, detail=f"document id={doc_id} not found")

    ok = await delete_doc_async(pool, doc_id)
    if not ok:
        raise HTTPException(status_code=404, detail=f"document id={doc_id} not found")

    # 清理上传的文件（只删 data/uploaded/ 下的，绝不动其他路径）
    source = doc.get("source_file")
    if source:
        p = _project_root() / str(source)
        try:
            if p.is_file() and p.parent == _upload_dir():
                p.unlink()
        except OSError:
            logger.warning("[upload] 清理源文件失败: %s", p)
    return {"deleted": doc_id}
