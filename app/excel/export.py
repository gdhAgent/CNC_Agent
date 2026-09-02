"""
app.excel.export —— 导出（按条件过滤 → xlsx，列结构与模板一致）

导出的文件改完能直接再导入。
"""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook

from app.excel.templates import get_columns


def export_alarms_to_xlsx(
    *,
    brand: str | None = None,
    origin: str | None = None,
) -> bytes:
    """
    导出报警码到 xlsx。
    列结构：与 alarm 模板一致（label + 说明行不写，避免循环）。
    """
    import psycopg

    from app.config import get_settings

    cfg = get_settings()
    sql = ["SELECT brand, controller, code, category, severity, name, "
           "       description, cause, action, safety_note, page_no "
           "  FROM kb.alarms WHERE 1=1"]
    params: list[Any] = []
    if brand:
        sql.append("AND brand = %s")
        params.append(brand)
    if origin:
        sql.append("AND origin = %s")
        params.append(origin)
    sql.append("ORDER BY brand, code_norm")

    cols = get_columns("alarm")

    with psycopg.connect(**cfg.db_dsn_kwargs()) as conn, conn.cursor() as cur:
        cur.execute(" ".join(sql), params)
        rows = cur.fetchall()

    return _build_xlsx(cols, rows, sheet_name="alarm")


def export_faq_to_xlsx(*, brand: str | None = None) -> bytes:
    """导出 FAQ 到 xlsx"""
    import psycopg

    from app.config import get_settings

    cfg = get_settings()
    sql = ["SELECT c.heading_path AS title, "
           "       COALESCE(d.brand, '') AS brand, "
           "       array_to_string(d.model_scope, ',') AS model_scope, "
           "       c.content AS body, "
           "       d.source_file AS source "
           "  FROM kb.chunks c "
           "  JOIN kb.documents d ON d.id = c.doc_id "
           " WHERE d.doc_type = 'faq' AND c.level = 2"]
    params: list[Any] = []
    if brand:
        sql.append("AND d.brand = %s")
        params.append(brand)
    sql.append("ORDER BY d.id")

    cols = get_columns("faq")
    with psycopg.connect(**cfg.db_dsn_kwargs()) as conn, conn.cursor() as cur:
        cur.execute(" ".join(sql), params)
        rows = cur.fetchall()

    return _build_xlsx(cols, rows, sheet_name="faq")


def _build_xlsx(
    cols: list[tuple[str, str, bool, str, list[str] | None]],
    rows: list[tuple],
    *,
    sheet_name: str,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = sheet_name

    # 第 1 行：label
    for ci, (_k, label, _r, _e, _enum) in enumerate(cols, start=1):
        ws.cell(row=1, column=ci, value=label)

    # 数据行：按 cols 顺序取值（rows 列顺序应与 cols 对齐）
    for ri, row in enumerate(rows, start=2):
        for ci, _ in enumerate(cols, start=1):
            if ci - 1 < len(row):
                ws.cell(row=ri, column=ci, value=row[ci - 1])
            else:
                ws.cell(row=ri, column=ci, value=None)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
