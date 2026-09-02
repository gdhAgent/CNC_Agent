"""
app.excel.templates —— Excel 模板元数据与生成。

四类模板：alarm / faq / machine / maintenance，含表头字段（key,label,required,example）、
枚举列下拉（DataValidation）、说明行与 2 行示例数据。

列结构约定被 import 解析 / export 输出共用：按此导出的文件改完能直接再导入。
"""

from __future__ import annotations

from io import BytesIO
from typing import Any, Literal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

TemplateType = Literal["alarm", "faq", "machine", "maintenance"]


# ===== 模板元数据 =====

_HEADER_FILL = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
_BOLD = Font(bold=True)
_NOTE_FILL = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")  # 浅黄


# 各模板的列定义
# (key, label, required, example, enum_values_or_None)
ALARM_COLUMNS: list[tuple[str, str, bool, str, list[str] | None]] = [
    ("brand", "品牌", True, "FANUC", ["FANUC", "MITSUBISHI", "SIEMENS", "HEIDENHAIN", "OTHER"]),
    ("controller", "系统", False, "0i-MF", None),
    ("code", "报警码", True, "SV0401", None),
    ("category", "类别", False, "servo",
     ["servo", "spindle", "pmc", "overtravel", "program", "system", "estop",
      "toolchange", "axis", "hydraulic", "pneumatic", "encoder", "param", "other", "unknown"]),
    ("severity", "严重度", False, "warning",
     ["info", "warning", "fault", "fatal", "unknown"]),
    ("name", "报警名称", True, "伺服放大器 VRDY 信号断开", None),
    ("description", "现象", False, "主轴电机...", None),
    ("cause", "可能原因（多行）", False, "原因1\n原因2", None),
    ("action", "处置步骤（多行）", False, "步骤A\n步骤B", None),
    ("safety_note", "安全提示", False, "断电后等待 5 分钟", None),
    ("page_no", "手册页码", False, "42", None),
]


FAQ_COLUMNS: list[tuple[str, str, bool, str, list[str] | None]] = [
    ("title", "标题", True, "主轴异响处理经验", None),
    ("brand", "品牌", False, "FANUC",
     ["FANUC", "MITSUBISHI", "SIEMENS", "HEIDENHAIN", "OTHER"]),
    ("model_scope", "适用机型（逗号分隔）", False, "VMC850,TC500", None),
    ("body", "正文", True, "常见故障：主轴异响。处置：...", None),
    ("source", "来源说明", False, "工程师 E1024 经验", None),
]


MACHINE_COLUMNS: list[tuple[str, str, bool, str, list[str] | None]] = [
    ("asset_no", "资产编号", True, "CN-001", None),
    ("name", "设备名称", True, "立式加工中心-03", None),
    ("brand", "品牌", True, "FANUC",
     ["FANUC", "MITSUBISHI", "SIEMENS", "HEIDENHAIN", "OTHER"]),
    ("model", "机型", False, "VMC850", None),
    ("controller", "系统", False, "0i-MF", None),
    ("workshop", "车间", False, "一车间", None),
    ("line_no", "产线", False, "L1", None),
    ("status", "状态", False, "running", ["running", "idle", "repair", "scrapped"]),
]


MAINTENANCE_COLUMNS: list[tuple[str, str, bool, str, list[str] | None]] = [
    ("order_no", "工单号", True, "WO-202601-00001", None),
    ("asset_no", "机台资产编号", True, "CN-001", None),
    ("alarm_code", "报警码", False, "SV0401", None),
    ("fault_type", "故障类型", False, "机械",
     ["机械", "电气", "液压", "气动", "软件"]),
    ("symptom", "现象描述", True, "加工件表面出现振纹", None),
    ("root_cause", "根本原因", False, "主轴电机...", None),
    ("action_taken", "处置过程", False, "清理风扇...", None),
    ("engineer", "维修工程师", False, "E1024", None),
    ("downtime_min", "停机时长（分钟）", False, "120", None),
    ("started_at", "开始时间 (ISO)", True, "2026-01-15T08:30:00+08:00", None),
    ("finished_at", "结束时间 (ISO)", False, "2026-01-15T10:30:00+08:00", None),
]


_TEMPLATE_REGISTRY: dict[str, list[tuple[str, str, bool, str, list[str] | None]]] = {
    "alarm": ALARM_COLUMNS,
    "faq": FAQ_COLUMNS,
    "machine": MACHINE_COLUMNS,
    "maintenance": MAINTENANCE_COLUMNS,
}


def get_columns(template_type: TemplateType) -> list[tuple[str, str, bool, str, list[str] | None]]:
    if template_type not in _TEMPLATE_REGISTRY:
        raise ValueError(f"unknown template type: {template_type!r}")
    return _TEMPLATE_REGISTRY[template_type]


# ===== 生成模板 =====

def generate_template_bytes(template_type: TemplateType) -> bytes:
    """
    生成模板 xlsx 的二进制内容。
    结构：
        第 1 行：表头（key + label）
        第 2 行：说明（"必填"标记 + 简短说明）
        第 3 行：示例 1
        第 4 行：示例 2
    """
    cols = get_columns(template_type)
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = template_type.upper()

    # 第 1 行：表头（key 用列字段名，label 是中文）
    for ci, (_key, label, _req, _ex, _enum) in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=ci, value=label)
        cell.font = _BOLD
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 第 2 行：说明行（"列名 | 必填 | 说明"）
    for ci, (key, _label, required, example, _enum) in enumerate(cols, start=1):
        note = f"key={key}"
        if required:
            note += " | 必填"
        if example:
            note += f" | 示例: {example}"
        cell = ws.cell(row=2, column=ci, value=note)
        cell.fill = _NOTE_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # 第 3~4 行：示例数据
    examples = _build_examples(template_type)
    for ri, row_data in enumerate(examples, start=3):
        for ci, (key, _label, _r, _e, _enum) in enumerate(cols, start=1):
            ws.cell(row=ri, column=ci, value=row_data.get(key, ""))

    # 列宽自适应
    for ci, (_key, label, _req, _ex, _e) in enumerate(cols, start=1):
        width = max(12, len(label) * 2 + 4)
        ws.column_dimensions[get_column_letter(ci)].width = width

    # 枚举列加 DataValidation
    for ci, (_key, _label, _req, _ex, enum_values) in enumerate(cols, start=1):
        if enum_values:
            dv = DataValidation(
                type="list",
                formula1='"' + ",".join(enum_values) + '"',
                allow_blank=True,
                showErrorMessage=True,
                errorTitle="非法值",
                error="请从下拉列表中选择，或留空（系统会归一为 unknown）",
            )
            ws.add_data_validation(dv)
            dv.add(f"{get_column_letter(ci)}3:{get_column_letter(ci)}1048576")

    # 顶部留 1 行空白（视觉缓冲）
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_examples(template_type: TemplateType) -> list[dict[str, Any]]:
    if template_type == "alarm":
        return [
            {
                "brand": "FANUC", "controller": "0i-MF", "code": "SV0401",
                "category": "servo", "severity": "fault",
                "name": "伺服放大器 VRDY 信号断开",
                "description": "上电后伺服就绪信号异常",
                "cause": "伺服放大器电源异常\n急停回路未复位",
                "action": "检查 POWER LED\n测量急停按钮\n更换伺服放大器",
                "safety_note": "断电后等待 5 分钟放电",
                "page_no": "42",
            },
            {
                "brand": "MITSUBISHI", "controller": "M70", "code": "AL24",
                "category": "servo", "severity": "warning",
                "name": "初始参数未设定",
                "description": "驱动器初次上电报警",
                "cause": "出厂参数未初始化",
                "action": "执行参数初始化步骤",
                "safety_note": "确保机械处于安全位置",
                "page_no": "12",
            },
        ]
    if template_type == "faq":
        return [
            {
                "title": "主轴异响处理经验",
                "brand": "FANUC", "model_scope": "VMC850,TC500",
                "body": "常见故障：主轴异响。处置：1. 检查刀具 2. 检查轴承 3. 润滑",
                "source": "工程师 E1024",
            },
            {
                "title": "换刀故障排查",
                "brand": "FANUC", "model_scope": "",
                "body": "刀库不转动时检查气压、电磁阀、机械臂位置",
                "source": "工程师 E1025",
            },
        ]
    if template_type == "machine":
        return [
            {
                "asset_no": "CN-001", "name": "立式加工中心-01",
                "brand": "FANUC", "model": "VMC850", "controller": "0i-MF",
                "workshop": "一车间", "line_no": "L1", "status": "running",
            },
            {
                "asset_no": "CN-002", "name": "车床-02",
                "brand": "MITSUBISHI", "model": "TC500", "controller": "M70",
                "workshop": "二车间", "line_no": "L1", "status": "idle",
            },
        ]
    if template_type == "maintenance":
        return [
            {
                "order_no": "WO-202601-00001", "asset_no": "CN-001",
                "alarm_code": "SV0401", "fault_type": "电气",
                "symptom": "上电报 SV0401", "root_cause": "急停按钮未复位",
                "action_taken": "复位急停后恢复", "engineer": "E1024",
                "downtime_min": "30", "started_at": "2026-01-15T08:30:00+08:00",
                "finished_at": "2026-01-15T09:00:00+08:00",
            },
        ]
    return []
