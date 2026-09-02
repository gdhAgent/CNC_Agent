"""
app.excel.validate —— Excel 导入行校验（validate + confirm 两阶段，供 /api/import 用）。

validate：解析上传 xlsx → 逐行校验（必填 / 枚举 / 时间格式等），返回 ValidationResult，
**不写入业务表**（查重在 confirm 阶段做）。confirm：按 job_id 与 dup_strategy 实际入库，
后台异步进行，进度写入 kb.import_jobs。
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from io import BytesIO
from typing import Any, Literal

from openpyxl import load_workbook

from app.excel.templates import get_columns

logger = logging.getLogger(__name__)


DupStrategy = Literal["skip", "overwrite", "duplicate"]


@dataclass(slots=True, frozen=True)
class RowError:
    """单条行的字段级错误"""
    line_no: int
    field: str | None
    reason: str


@dataclass(slots=True)
class ValidationResult:
    """一次 validate 的完整结果（写入 kb.import_jobs.errors 字段用）"""
    total_rows: int = 0
    valid_rows: int = 0
    dup_rows: int = 0      # 重复（命中 unique key）的行数
    error_rows: int = 0    # 校验失败的行数
    errors: list[RowError] = field(default_factory=list)
    # 解析出的原始 rows（含字段值），供 confirm 阶段直接消费
    parsed_rows: list[dict[str, Any]] = field(default_factory=list)


# ===== 工具函数 =====

def _norm_cell(value: Any) -> str | None:
    """Excel 单元格 → str（去空白，None 视为空）"""
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def parse_uploaded_xlsx(content: bytes, template_type: str) -> list[tuple[int, dict[str, Any]]]:
    """
    解析上传的 xlsx → [(line_no, row_dict), ...]
    line_no 从 2 起（header 是第 1 行）。
    空行跳过。
    """
    cols = get_columns(template_type)  # type: ignore[arg-type]
    keys = [c[0] for c in cols]

    wb = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    assert ws is not None

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return []

    header_list = [_norm_cell(h) for h in header]
    if not header_list or header_list[0] != cols[0][1]:
        # 表头第 1 列应是 label，可能客户端改了顺序；改为按 label 匹配
        # 这里采用保守策略：要求 header 第 1 列等于第 1 个 label，否则当作空文件
        first_label = header_list[0] if header_list else None
        logger.warning(
            "[excel] header mismatch; expected %r got %r", cols[0][1], first_label
        )
        return []

    result: list[tuple[int, dict[str, Any]]] = []
    for ln, row in enumerate(rows_iter, start=2):
        # 跳过空行
        if row is None or all(_norm_cell(c) is None for c in row):
            continue
        # header_list 可能比 keys 长/短；按 index 对齐
        rec: dict[str, Any] = {}
        for ci, key in enumerate(keys):
            if ci < len(header_list):
                rec[key] = _norm_cell(row[ci]) if ci < len(row) else None
            else:
                rec[key] = None
        result.append((ln, rec))
    return result


# ===== 字段校验 =====

_ISO_DT = re.compile(r"^\d{4}-\d{2}-\d{2}(T\d{2}:\d{2}:\d{2}(\+\d{2}:\d{2})?)?$")


def _check_alarm_row(rec: dict[str, Any]) -> list[RowError]:
    errs: list[RowError] = []
    line = rec.get("__line_no__", -1)

    # 必填：brand / code / name
    for f in ("brand", "code", "name"):
        if not rec.get(f):
            errs.append(RowError(line, f, "必填字段为空"))

    # 枚举：category / severity
    enum_cat = next((c[4] for c in get_columns("alarm") if c[0] == "category"), None)
    enum_sev = next((c[4] for c in get_columns("alarm") if c[0] == "severity"), None)
    if rec.get("category") and enum_cat and rec["category"] not in enum_cat:
        errs.append(RowError(line, "category", f"不在枚举 {enum_cat} 内（未知值会归一为 unknown）"))
    if rec.get("severity") and enum_sev and rec["severity"] not in enum_sev:
        errs.append(RowError(line, "severity", f"不在枚举 {enum_sev} 内（未知值会归一为 unknown）"))

    # code 格式：3~6 位数字 + 可选 2 字母前缀
    if rec.get("code") and not re.match(r"^(?:[A-Z]{2})?\d{2,6}$", rec["code"].upper()):
        errs.append(RowError(line, "code", "格式不合法（应如 SV0401 / 3001）"))

    # page_no：可选；填了就要是整数
    if rec.get("page_no"):
        try:
            int(rec["page_no"])
        except ValueError:
            errs.append(RowError(line, "page_no", "应填整数"))

    return errs


def _check_faq_row(rec: dict[str, Any]) -> list[RowError]:
    errs: list[RowError] = []
    line = rec.get("__line_no__", -1)
    for fld in ("title", "body"):
        if not rec.get(fld):
            errs.append(RowError(line, fld, "必填字段为空"))
    return errs


def _check_machine_row(rec: dict[str, Any]) -> list[RowError]:
    errs: list[RowError] = []
    line = rec.get("__line_no__", -1)
    for fld in ("asset_no", "name", "brand"):
        if not rec.get(fld):
            errs.append(RowError(line, fld, "必填字段为空"))
    cols = get_columns("machine")
    enum_status = next((c[4] for c in cols if c[0] == "status"), None)
    if rec.get("status") and enum_status and rec["status"] not in enum_status:
        errs.append(RowError(line, "status", f"不在枚举 {enum_status} 内"))
    return errs


def _check_maintenance_row(rec: dict[str, Any]) -> list[RowError]:
    errs: list[RowError] = []
    line = rec.get("__line_no__", -1)
    for fld in ("order_no", "asset_no", "symptom", "started_at"):
        if not rec.get(fld):
            errs.append(RowError(line, fld, "必填字段为空"))
    for fld in ("started_at", "finished_at"):
        if rec.get(fld) and not _ISO_DT.match(rec[fld]):
            errs.append(RowError(line, fld, "应填 ISO 时间 (YYYY-MM-DD 或 YYYY-MM-DDTHH:MM:SS+TZ)"))
    if rec.get("downtime_min"):
        try:
            int(rec["downtime_min"])
        except ValueError:
            errs.append(RowError(line, "downtime_min", "应填整数（分钟）"))
    return errs


_ROW_CHECKERS = {
    "alarm": _check_alarm_row,
    "faq": _check_faq_row,
    "machine": _check_machine_row,
    "maintenance": _check_maintenance_row,
}


def validate_rows(template_type: str, rows: list[tuple[int, dict[str, Any]]]) -> ValidationResult:
    """
    校验已解析的行。返回 ValidationResult（含 errors 和 parsed_rows）。
    不做 DB 查重 —— 那在 confirm 阶段做。
    """
    result = ValidationResult()
    result.total_rows = len(rows)

    checker = _ROW_CHECKERS.get(template_type)
    if checker is None:
        return result

    for line_no, row in rows:
        row["__line_no__"] = line_no
        errs = checker(row)
        if errs:
            result.error_rows += 1
            result.errors.extend(errs)
        else:
            result.valid_rows += 1
        result.parsed_rows.append(row)
    return result


# ===== 顶层入口 =====

def parse_and_validate(content: bytes, template_type: str) -> ValidationResult:
    """上传 xlsx → 解析 → 校验 → 一次性返回结果"""
    rows = parse_uploaded_xlsx(content, template_type)
    return validate_rows(template_type, rows)
